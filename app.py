from __future__ import annotations

import csv
import html
import json
import os
import queue
import re
import threading
import time
import traceback
import urllib.parse
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError as exc:
    raise SystemExit("当前 Python 环境缺少 Tkinter，无法启动图形界面。") from exc

APP_NAME = "DOI 文献批量下载器"
APP_VERSION = "1.0.0"
USER_AGENT = f"paper-downloader/{APP_VERSION} (open-access literature downloader)"
DOI_PATTERN = re.compile(r"10\.\d{4,9}/[-._;()/:A-Z0-9]+", re.IGNORECASE)
NUMBER_PATTERN = re.compile(
    r"(?m)^\s*(?:\[(?P<bracket>\d{1,6})\]|(?P<plain>\d{1,6})\s*[\.．、\)])\s*"
)
PDF_META_PATTERNS = [
    re.compile(
        r'<meta[^>]+name=["\']citation_pdf_url["\'][^>]+content=["\']([^"\']+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']citation_pdf_url["\']',
        re.IGNORECASE,
    ),
]
ILLEGAL_FILENAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
TRAILING_DOI_PUNCTUATION = ".,;:!?)]}>\"'，。；：！？）】》」』"


@dataclass(frozen=True)
class ReferenceRecord:
    number: str
    doi: str
    raw_reference: str


@dataclass
class DownloadResult:
    number: str
    doi: str
    status: str
    filename: str = ""
    source: str = ""
    url: str = ""
    message: str = ""
    raw_reference: str = ""
    elapsed_seconds: float = 0.0


@dataclass(frozen=True)
class CandidateUrl:
    url: str
    source: str


def normalize_doi(value: str) -> str:
    value = html.unescape(urllib.parse.unquote(value)).strip()
    value = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", value, flags=re.IGNORECASE)
    value = re.sub(r"^doi\s*:\s*", "", value, flags=re.IGNORECASE)
    value = value.strip().rstrip(TRAILING_DOI_PUNCTUATION)
    value = value.split("#", 1)[0].split("?", 1)[0]
    return value.lower()


def extract_dois(text: str) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    for match in DOI_PATTERN.finditer(text):
        doi = normalize_doi(match.group(0))
        if doi and doi not in seen:
            seen.add(doi)
            found.append(doi)
    return found


def parse_references(text: str) -> list[ReferenceRecord]:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    matches = list(NUMBER_PATTERN.finditer(normalized))
    records: list[ReferenceRecord] = []
    seen_dois: set[str] = set()

    if matches:
        for index, match in enumerate(matches):
            number = match.group("bracket") or match.group("plain") or str(index + 1)
            start = match.start()
            end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
            block = normalized[start:end].strip()
            dois = extract_dois(block)
            for doi_index, doi in enumerate(dois, start=1):
                if doi in seen_dois:
                    continue
                seen_dois.add(doi)
                effective_number = number if doi_index == 1 else f"{number}-{doi_index}"
                records.append(ReferenceRecord(effective_number, doi, block))
    else:
        for index, doi in enumerate(extract_dois(normalized), start=1):
            records.append(ReferenceRecord(str(index), doi, normalized.strip()))

    return records


def find_numbered_entries_without_doi(text: str) -> list[DownloadResult]:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    matches = list(NUMBER_PATTERN.finditer(normalized))
    missing: list[DownloadResult] = []
    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
        block = normalized[start:end].strip()
        if extract_dois(block):
            continue
        number = match.group("bracket") or match.group("plain") or str(index + 1)
        missing.append(
            DownloadResult(
                number=number,
                doi="",
                status="无有效DOI",
                message="该编号条目中未识别到完整 DOI",
                raw_reference=block,
            )
        )
    return missing


def read_text_auto(path: Path) -> str:
    raw = path.read_bytes()
    encodings = ("utf-8-sig", "utf-16", "gb18030", "big5")
    errors: list[str] = []
    for encoding in encodings:
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError as exc:
            errors.append(f"{encoding}: {exc}")
    raise UnicodeError("无法识别 TXT 编码。已尝试：" + "; ".join(errors))


def safe_filename_component(value: str) -> str:
    cleaned = ILLEGAL_FILENAME.sub("_", value)
    cleaned = re.sub(r"\s+", "_", cleaned).strip(" ._")
    return cleaned or "unknown"


def build_pdf_filename(number: str, doi: str) -> str:
    return f"{safe_filename_component(number)}+{safe_filename_component(doi)}.pdf"


class OpenAccessResolver:
    def __init__(self, email: str, timeout: int = 25) -> None:
        self.email = email.strip()
        self.timeout = timeout
        self.session = requests.Session()
        self.session.trust_env = False
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/pdf,text/html,application/json;q=0.9,*/*;q=0.8",
            }
        )

    def _json_get(self, url: str, params: Optional[dict[str, str]] = None) -> dict:
        response = self.session.get(url, params=params, timeout=self.timeout)
        response.raise_for_status()
        return response.json()

    def candidates(self, doi: str) -> list[CandidateUrl]:
        candidates: list[CandidateUrl] = []
        seen: set[str] = set()

        def add(url: Optional[str], source: str) -> None:
            if not url:
                return
            url = urllib.parse.urljoin(f"https://doi.org/{doi}", url.strip())
            if url.startswith(("http://", "https://")) and url not in seen:
                seen.add(url)
                candidates.append(CandidateUrl(url, source))

        if self.email:
            try:
                payload = self._json_get(
                    f"https://api.unpaywall.org/v2/{urllib.parse.quote(doi, safe='')}",
                    params={"email": self.email},
                )
                best = payload.get("best_oa_location") or {}
                add(best.get("url_for_pdf"), "Unpaywall")
                add(best.get("url"), "Unpaywall")
                for location in payload.get("oa_locations") or []:
                    add(location.get("url_for_pdf"), "Unpaywall")
                    add(location.get("url"), "Unpaywall")
            except (requests.RequestException, ValueError, TypeError):
                pass

        try:
            payload = self._json_get(
                "https://api.openalex.org/works/" + urllib.parse.quote(f"https://doi.org/{doi}", safe="")
            )
            best = payload.get("best_oa_location") or {}
            primary = payload.get("primary_location") or {}
            add(best.get("pdf_url"), "OpenAlex")
            add(best.get("landing_page_url"), "OpenAlex")
            add(primary.get("pdf_url"), "OpenAlex")
            add(primary.get("landing_page_url"), "OpenAlex")
            for location in payload.get("locations") or []:
                add(location.get("pdf_url"), "OpenAlex")
                add(location.get("landing_page_url"), "OpenAlex")
        except (requests.RequestException, ValueError, TypeError):
            pass

        try:
            payload = self._json_get(
                f"https://api.crossref.org/works/{urllib.parse.quote(doi, safe='')}"
            )
            message = payload.get("message") or {}
            for link in message.get("link") or []:
                content_type = str(link.get("content-type") or "").lower()
                if "pdf" in content_type or "unspecified" in content_type:
                    add(link.get("URL"), "Crossref")
        except (requests.RequestException, ValueError, TypeError):
            pass

        add(f"https://doi.org/{doi}", "DOI 页面")
        return candidates

    def _extract_pdf_links(self, response: requests.Response) -> list[str]:
        content_type = response.headers.get("Content-Type", "").lower()
        if "html" not in content_type and not response.text.lstrip().startswith("<"):
            return []
        text = response.text[:2_000_000]
        urls: list[str] = []
        seen: set[str] = set()
        for pattern in PDF_META_PATTERNS:
            for match in pattern.finditer(text):
                url = html.unescape(match.group(1).strip())
                absolute = urllib.parse.urljoin(response.url, url)
                if absolute not in seen:
                    seen.add(absolute)
                    urls.append(absolute)
        for match in re.finditer(r'href=["\']([^"\']+\.pdf(?:\?[^"\']*)?)["\']', text, re.IGNORECASE):
            absolute = urllib.parse.urljoin(response.url, html.unescape(match.group(1)))
            if absolute not in seen:
                seen.add(absolute)
                urls.append(absolute)
        return urls

    def _download_one_url(
        self,
        candidate: CandidateUrl,
        target: Path,
        visited: Optional[set[str]] = None,
        depth: int = 0,
    ) -> tuple[bool, str, str]:
        visited = visited or set()
        if candidate.url in visited:
            return False, candidate.url, "检测到重复跳转链接"
        if depth > 2:
            return False, candidate.url, "网页内 PDF 跳转层级过深"
        visited.add(candidate.url)
        try:
            response = self.session.get(
                candidate.url,
                timeout=self.timeout,
                allow_redirects=True,
                stream=True,
            )
            response.raise_for_status()
            first_chunk = next(response.iter_content(chunk_size=64 * 1024), b"")

            if b"%PDF-" in first_chunk[:1024]:
                temp = target.with_suffix(target.suffix + ".part")
                with temp.open("wb") as file_handle:
                    file_handle.write(first_chunk)
                    for chunk in response.iter_content(chunk_size=256 * 1024):
                        if chunk:
                            file_handle.write(chunk)
                if temp.stat().st_size < 1024:
                    temp.unlink(missing_ok=True)
                    return False, response.url, "PDF 文件过小，已拒绝保存"
                temp.replace(target)
                return True, response.url, ""

            response.close()
            html_response = self.session.get(candidate.url, timeout=self.timeout, allow_redirects=True)
            html_response.raise_for_status()
            for pdf_url in self._extract_pdf_links(html_response):
                nested = CandidateUrl(pdf_url, candidate.source + " 页面")
                success, final_url, message = self._download_one_url(
                    nested, target, visited=visited, depth=depth + 1
                )
                if success:
                    return success, final_url, message
            return False, html_response.url, "链接未返回 PDF"
        except requests.RequestException as exc:
            return False, candidate.url, f"网络错误：{exc}"
        except OSError as exc:
            return False, candidate.url, f"文件写入错误：{exc}"

    def download(self, record: ReferenceRecord, output_dir: Path, cancel_event: threading.Event) -> DownloadResult:
        started = time.perf_counter()
        filename = build_pdf_filename(record.number, record.doi)
        target = output_dir / filename

        if cancel_event.is_set():
            return DownloadResult(
                record.number,
                record.doi,
                "已取消",
                filename=filename,
                message="任务已取消",
                raw_reference=record.raw_reference,
            )

        if target.exists() and target.stat().st_size >= 1024:
            return DownloadResult(
                record.number,
                record.doi,
                "已存在",
                filename=filename,
                message="目标文件已存在，未重复下载",
                raw_reference=record.raw_reference,
                elapsed_seconds=round(time.perf_counter() - started, 2),
            )

        candidates = self.candidates(record.doi)
        last_message = "未找到开放获取 PDF"
        last_url = ""
        for candidate in candidates:
            if cancel_event.is_set():
                return DownloadResult(
                    record.number,
                    record.doi,
                    "已取消",
                    filename=filename,
                    message="任务已取消",
                    raw_reference=record.raw_reference,
                    elapsed_seconds=round(time.perf_counter() - started, 2),
                )
            success, final_url, message = self._download_one_url(candidate, target)
            last_message = message or last_message
            last_url = final_url
            if success:
                return DownloadResult(
                    record.number,
                    record.doi,
                    "下载成功",
                    filename=filename,
                    source=candidate.source,
                    url=final_url,
                    raw_reference=record.raw_reference,
                    elapsed_seconds=round(time.perf_counter() - started, 2),
                )

        return DownloadResult(
            record.number,
            record.doi,
            "下载失败",
            filename=filename,
            url=last_url,
            message=last_message,
            raw_reference=record.raw_reference,
            elapsed_seconds=round(time.perf_counter() - started, 2),
        )


def write_logs(output_dir: Path, results: Iterable[DownloadResult]) -> tuple[Path, Path]:
    rows = list(results)
    xlsx_path = output_dir / "下载日志.xlsx"
    csv_path = output_dir / "下载日志.csv"
    headers = [
        "编号",
        "DOI",
        "状态",
        "文件名",
        "下载来源",
        "最终网址",
        "说明",
        "耗时（秒）",
        "原始文献条目",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "下载日志"
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    failure_fill = PatternFill("solid", fgColor="FFF2CC")
    success_fill = PatternFill("solid", fgColor="E2F0D9")
    canceled_fill = PatternFill("solid", fgColor="FCE4D6")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for result in rows:
        sheet.append(
            [
                result.number,
                result.doi,
                result.status,
                result.filename,
                result.source,
                result.url,
                result.message,
                result.elapsed_seconds,
                result.raw_reference,
            ]
        )
        row_index = sheet.max_row
        if result.status == "下载成功":
            fill = success_fill
        elif result.status in {"下载失败", "无有效DOI"}:
            fill = failure_fill
        elif result.status == "已取消":
            fill = canceled_fill
        else:
            fill = None
        if fill:
            for cell in sheet[row_index]:
                cell.fill = fill
        for cell in sheet[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [10, 35, 14, 55, 18, 60, 45, 12, 90]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(xlsx_path)

    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for result in rows:
            writer.writerow(
                {
                    "编号": result.number,
                    "DOI": result.doi,
                    "状态": result.status,
                    "文件名": result.filename,
                    "下载来源": result.source,
                    "最终网址": result.url,
                    "说明": result.message,
                    "耗时（秒）": result.elapsed_seconds,
                    "原始文献条目": result.raw_reference,
                }
            )
    return xlsx_path, csv_path


class PaperDownloaderApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"{APP_NAME} v{APP_VERSION}")
        self.geometry("920x680")
        self.minsize(820, 600)
        self.cancel_event = threading.Event()
        self.event_queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self.worker_thread: Optional[threading.Thread] = None
        self.results: list[DownloadResult] = []

        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar(value=str(Path.home() / "Downloads" / "DOI文献下载"))
        self.email = tk.StringVar()
        self.worker_count = tk.IntVar(value=4)
        self.status_text = tk.StringVar(value="请选择包含编号和 DOI 的 TXT 文档。")
        self.progress_value = tk.DoubleVar(value=0)

        self._build_ui()
        self.after(150, self._poll_events)

    def _build_ui(self) -> None:
        root = ttk.Frame(self, padding=16)
        root.pack(fill="both", expand=True)
        root.columnconfigure(1, weight=1)
        root.rowconfigure(6, weight=1)

        ttk.Label(root, text="TXT 文档：").grid(row=0, column=0, sticky="w", pady=6)
        ttk.Entry(root, textvariable=self.input_path).grid(row=0, column=1, sticky="ew", padx=8)
        ttk.Button(root, text="选择文件", command=self._choose_input).grid(row=0, column=2)

        ttk.Label(root, text="下载目录：").grid(row=1, column=0, sticky="w", pady=6)
        ttk.Entry(root, textvariable=self.output_path).grid(row=1, column=1, sticky="ew", padx=8)
        ttk.Button(root, text="选择目录", command=self._choose_output).grid(row=1, column=2)

        ttk.Label(root, text="联系邮箱：").grid(row=2, column=0, sticky="w", pady=6)
        ttk.Entry(root, textvariable=self.email).grid(row=2, column=1, sticky="ew", padx=8)
        ttk.Label(root, text="用于 Unpaywall 合规查询，建议填写真实邮箱").grid(row=2, column=2, sticky="w")

        ttk.Label(root, text="并发数量：").grid(row=3, column=0, sticky="w", pady=6)
        ttk.Spinbox(root, from_=1, to=8, textvariable=self.worker_count, width=8).grid(
            row=3, column=1, sticky="w", padx=8
        )

        button_row = ttk.Frame(root)
        button_row.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(12, 8))
        self.start_button = ttk.Button(button_row, text="开始识别并下载", command=self._start)
        self.start_button.pack(side="left")
        self.cancel_button = ttk.Button(button_row, text="取消", command=self._cancel, state="disabled")
        self.cancel_button.pack(side="left", padx=8)
        ttk.Button(button_row, text="打开下载目录", command=self._open_output).pack(side="left")
        ttk.Button(button_row, text="项目主页", command=lambda: webbrowser.open("https://github.com/luxiaoshi838-jpg/paper-downloader")).pack(side="right")

        progress = ttk.Progressbar(root, variable=self.progress_value, maximum=100)
        progress.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(2, 6))

        log_frame = ttk.LabelFrame(root, text="运行记录", padding=8)
        log_frame.grid(row=6, column=0, columnspan=3, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.log_box = tk.Text(log_frame, wrap="word", state="disabled")
        self.log_box.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_box.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.log_box.configure(yscrollcommand=scrollbar.set)

        ttk.Label(root, textvariable=self.status_text, anchor="w").grid(
            row=7, column=0, columnspan=3, sticky="ew", pady=(8, 0)
        )
        ttk.Label(
            root,
            text="仅检索和下载公开可访问或开放获取的 PDF，不绕过登录、付费墙或访问控制。",
            foreground="#666666",
        ).grid(row=8, column=0, columnspan=3, sticky="w", pady=(4, 0))

    def _choose_input(self) -> None:
        path = filedialog.askopenfilename(
            title="选择 TXT 文档",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
        )
        if path:
            self.input_path.set(path)
            if not self.output_path.get().strip():
                self.output_path.set(str(Path(path).with_name("下载文献")))

    def _choose_output(self) -> None:
        path = filedialog.askdirectory(title="选择下载目录")
        if path:
            self.output_path.set(path)

    def _open_output(self) -> None:
        path = Path(self.output_path.get().strip()).expanduser()
        path.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(path)
        except AttributeError:
            webbrowser.open(path.as_uri())

    def _append_log(self, text: str) -> None:
        self.log_box.configure(state="normal")
        self.log_box.insert("end", text.rstrip() + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _set_running(self, running: bool) -> None:
        self.start_button.configure(state="disabled" if running else "normal")
        self.cancel_button.configure(state="normal" if running else "disabled")

    def _start(self) -> None:
        input_value = self.input_path.get().strip()
        output_value = self.output_path.get().strip()
        if not input_value:
            messagebox.showwarning(APP_NAME, "请先选择 TXT 文档。")
            return
        input_path = Path(input_value)
        if not input_path.is_file():
            messagebox.showerror(APP_NAME, "选择的 TXT 文档不存在。")
            return
        if not output_value:
            messagebox.showwarning(APP_NAME, "请选择下载目录。")
            return

        workers = max(1, min(8, int(self.worker_count.get())))
        output_dir = Path(output_value).expanduser()
        output_dir.mkdir(parents=True, exist_ok=True)
        self.cancel_event.clear()
        self.results.clear()
        self.progress_value.set(0)
        self._set_running(True)
        self._append_log("=" * 60)
        self._append_log(f"读取文档：{input_path}")
        self.status_text.set("正在识别 DOI……")
        self.worker_thread = threading.Thread(
            target=self._run_download,
            args=(input_path, output_dir, self.email.get(), workers),
            daemon=True,
        )
        self.worker_thread.start()

    def _cancel(self) -> None:
        self.cancel_event.set()
        self.status_text.set("正在取消尚未开始的任务……")
        self._append_log("已请求取消。正在运行的网络请求完成后将停止。")

    def _run_download(self, input_path: Path, output_dir: Path, email: str, workers: int) -> None:
        try:
            text = read_text_auto(input_path)
            records = parse_references(text)
            missing_results = find_numbered_entries_without_doi(text)
            if not records and not missing_results:
                self.event_queue.put(("error", "未识别到文献编号或有效 DOI。请确认文档包含形如 10.xxxx/xxxxx 的 DOI。"))
                return
            total_items = len(records) + len(missing_results)
            self.event_queue.put(("parsed", (len(records), len(missing_results))))
            results: list[DownloadResult] = list(missing_results)
            for missing_index, missing in enumerate(missing_results, start=1):
                self.event_queue.put(("result", (missing, missing_index, total_items)))
            with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="paper-download") as executor:
                future_map = {
                    executor.submit(
                        OpenAccessResolver(email=email).download,
                        record,
                        output_dir,
                        self.cancel_event,
                    ): record
                    for record in records
                }
                completed = len(missing_results)
                for future in as_completed(future_map):
                    record = future_map[future]
                    try:
                        result = future.result()
                    except Exception as exc:
                        result = DownloadResult(
                            record.number,
                            record.doi,
                            "下载失败",
                            filename=build_pdf_filename(record.number, record.doi),
                            message=f"未处理异常：{exc}",
                            raw_reference=record.raw_reference,
                        )
                    results.append(result)
                    completed += 1
                    self.event_queue.put(("result", (result, completed, total_items)))
            results.sort(key=lambda item: natural_sort_key(item.number))
            xlsx_path, csv_path = write_logs(output_dir, results)
            self.event_queue.put(("finished", (results, xlsx_path, csv_path)))
        except Exception as exc:
            detail = "".join(traceback.format_exception_only(type(exc), exc)).strip()
            self.event_queue.put(("error", detail))

    def _poll_events(self) -> None:
        try:
            while True:
                event, payload = self.event_queue.get_nowait()
                if event == "parsed":
                    doi_count, missing_count = payload
                    self.status_text.set(f"已识别 {doi_count} 个不重复 DOI，开始下载……")
                    self._append_log(f"识别到 {doi_count} 个不重复 DOI。")
                    if missing_count:
                        self._append_log(f"另有 {missing_count} 个编号条目未识别到 DOI，已写入黄色日志。")
                elif event == "result":
                    result, completed, total = payload
                    self.results.append(result)
                    self.progress_value.set(completed / total * 100)
                    extra = f"；{result.message}" if result.message else ""
                    self._append_log(
                        f"[{completed}/{total}] {result.number} | {result.doi} | {result.status}{extra}"
                    )
                    self.status_text.set(f"已完成 {completed}/{total}")
                elif event == "finished":
                    results, xlsx_path, csv_path = payload
                    success = sum(item.status == "下载成功" for item in results)
                    existing = sum(item.status == "已存在" for item in results)
                    failed = sum(item.status in {"下载失败", "无有效DOI"} for item in results)
                    canceled = sum(item.status == "已取消" for item in results)
                    self.progress_value.set(100)
                    self.status_text.set(
                        f"完成：成功 {success}，已存在 {existing}，失败 {failed}，取消 {canceled}"
                    )
                    self._append_log(f"Excel 日志：{xlsx_path}")
                    self._append_log(f"CSV 日志：{csv_path}")
                    self._set_running(False)
                    messagebox.showinfo(
                        APP_NAME,
                        f"处理完成。\n\n成功：{success}\n已存在：{existing}\n失败：{failed}\n取消：{canceled}\n\n失败记录已在 Excel 日志中标黄。",
                    )
                elif event == "error":
                    self._set_running(False)
                    self.status_text.set("任务失败")
                    self._append_log(f"错误：{payload}")
                    messagebox.showerror(APP_NAME, str(payload))
        except queue.Empty:
            pass
        self.after(150, self._poll_events)


def natural_sort_key(value: str) -> tuple:
    return tuple(int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", value))


def main() -> None:
    app = PaperDownloaderApp()
    app.mainloop()


if __name__ == "__main__":
    main()
